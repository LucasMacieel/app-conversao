import math

import numpy as np
from openpyxl import load_workbook

from integracao.definicoes.definicoes import Transformador, Lamina, LaminaInformacoes, TransformadorInformacoes, \
    dadosMagnetizacao
from integracao.modulos.suporte import encontrar_produto_mais_proximo, encontrar_densidade_corrente, \
    encontrar_valor_awg, calcular_secao_cobre, calcular_densidade_corrente_media, calcular_espiras, verificar_valores,\
    custom_interpolation


def dimensionar_transformador(tensao_primaria, tensao_secundaria, potencia_secundaria, frequencia, espessura_lamina):
    transformador = Transformador.P1_S1
    lamina = Lamina.PADRONIZADA

    i1, i2 = calcular_correntes(potencia_secundaria, tensao_primaria, tensao_secundaria)
    densidade_corrente = encontrar_densidade_corrente(potencia_secundaria)

    if not verificar_valores(densidade_corrente):
        return False

    s1, s2 = calcular_secao_condutor(i1, i2, densidade_corrente)
    fio1, fio2, s1, s2 = calcular_bitolas(s1, s2)

    if not verificar_valores(fio1, fio2, s1, s2):
        return False

    densidade_corrente_media = calcular_densidade_corrente_media(i1, i2, s1, s2)

    aux_result = calcular_parametros_lamina(frequencia, lamina, potencia_secundaria, tensao_primaria,
                                            tensao_secundaria, transformador)

    if aux_result == "Inválido":
        return False

    a, b, n1, n2, sm = aux_result

    if a < 0 or b < 0 or not possibilidade_execucao(lamina, a, n1, n2, s1, s2):
        lamina = Lamina.COMPRIDA
        aux_result = calcular_parametros_lamina(frequencia, lamina, potencia_secundaria, tensao_primaria,
                                                tensao_secundaria, transformador)

        if (aux_result == "Inválido") or \
                (a < 0 or b < 0 or not possibilidade_execucao(lamina, a, n1, n2, s1, s2)):
            return False

        a, b, n1, n2, sm = aux_result

    quantidade_laminas = calcular_quantidade_laminas(b, espessura_lamina)
    p_fe = calcular_peso_ferro(lamina, a, b)

    if not verificar_valores(p_fe):
        return False

    s_cu = calcular_secao_cobre(n1, n2, s1, s2)
    p_cu = calcular_peso_cobre(a, b, s_cu)
    w_fe = calcular_perdas_ferro(frequencia, p_fe)
    w_cu = calcular_perdas_cobre(densidade_corrente_media, p_cu)
    rendimento = calcular_rendimento(potencia_secundaria, w_fe, w_cu)

    especifacacoes_transformador(a, b, fio1, fio2, lamina, n1, n2, p_cu, p_fe, quantidade_laminas, rendimento,
                                 transformador, w_cu, w_fe)

    return True


def especifacacoes_transformador(a, b, fio1, fio2, lamina, n1, n2, p_cu, p_fe, quantidade_laminas, rendimento,
                                 transformador, w_cu, w_fe):
    TransformadorInformacoes["tipo"] = transformador.name
    TransformadorInformacoes["lamina"] = lamina.name
    TransformadorInformacoes["quantidade_laminas"] = quantidade_laminas
    TransformadorInformacoes["espiras_primario"] = n1
    TransformadorInformacoes["espiras_secundario"] = n2
    TransformadorInformacoes["cabo_AWG_primario"] = fio1
    TransformadorInformacoes["cabo_AWG_secundario"] = fio2
    TransformadorInformacoes["dimensao_a"] = a
    TransformadorInformacoes["dimensao_b"] = b
    TransformadorInformacoes["peso_ferro"] = p_fe
    TransformadorInformacoes["peso_cobre"] = p_cu
    TransformadorInformacoes["peso_total"] = p_fe + p_cu
    TransformadorInformacoes["perdas_ferro"] = w_fe
    TransformadorInformacoes["perdas_cobre"] = w_cu
    TransformadorInformacoes["rendimento"] = rendimento


def calcular_parametros_lamina(frequencia, lamina, potencia_secundaria, tensao_primaria, tensao_secundaria,
                               transformador):
    sm = calcular_secao_magnetica(transformador, lamina, potencia_secundaria, frequencia)

    if not verificar_valores(sm):
        return "Inválido"

    a, b = dimensoes_nucleo(sm, lamina)

    if not verificar_valores(a, b):
        return "Inválido"

    sm = (a * b) / 1.1
    n1, n2 = numero_espiras(frequencia, tensao_primaria, tensao_secundaria, sm)

    return a, b, n1, n2, sm


def dimensoes_nucleo(secao_magnetica, tipo_lamina):
    secao_geometrica = secao_magnetica * 1.1

    dimensao = encontrar_produto_mais_proximo(secao_geometrica, tipo_lamina)
    if dimensao != "Inválido":
        dimensao_a, dimensao_b = dimensao
    else:
        return "Inválido", "Inválido"

    return dimensao_a, dimensao_b


def calcular_secao_magnetica(transformador, lamina, potencia, frequencia):
    secao_magnetica = 0

    if transformador == Transformador.P1_S1:
        if lamina == Lamina.PADRONIZADA:
            secao_magnetica = 7.5 * math.sqrt(potencia / frequencia)
        elif lamina == Lamina.COMPRIDA:
            secao_magnetica = 6 * math.sqrt(potencia / frequencia)
        elif lamina == Lamina.INVALIDO:
            return lamina.value
    elif transformador == Transformador.P2_S1_OU_P1_S2:
        if lamina == Lamina.PADRONIZADA:
            secao_magnetica = 7.5 * math.sqrt((1.25 * potencia) / frequencia)
        elif lamina == Lamina.COMPRIDA:
            secao_magnetica = 6 * math.sqrt((1.25 * potencia) / frequencia)
        elif lamina == Lamina.INVALIDO:
            return lamina.value
    elif transformador == Transformador.P2_S2:
        if lamina == Lamina.PADRONIZADA:
            secao_magnetica = 7.5 * math.sqrt((1.5 * potencia) / frequencia)
        elif lamina == Lamina.COMPRIDA:
            secao_magnetica = 6 * math.sqrt((1.5 * potencia) / frequencia)
        elif lamina == Lamina.INVALIDO:
            return lamina.value
    elif transformador == Transformador.INVALIDO:
        return "Inválido"

    return secao_magnetica


def calcular_quantidade_laminas(dimensao_b, espessura_lamina=0.5):
    return (dimensao_b * 0.9) / espessura_lamina


def numero_espiras(frequencia, tensao_primaria, tensao_secundaria, secao_magnetica, inducao_magnetica_max=11300):
    n_primario = calcular_espiras(tensao_primaria, inducao_magnetica_max, secao_magnetica, frequencia)
    n_secundario = calcular_espiras(tensao_secundaria, inducao_magnetica_max, secao_magnetica, frequencia)
    # Fator de 10% aplicado para compensar quedas de tensão
    n_secundario = math.ceil(1.1 * n_secundario)

    return n_primario, n_secundario


def calcular_bitolas(secao_condutor_primario, secao_condutor_secundario):
    awg_cabo_primario, secao_condutor_primario = encontrar_valor_awg(secao_condutor_primario)
    awg_cabo_secundario, secao_condutor_secundario = encontrar_valor_awg(secao_condutor_secundario)

    return awg_cabo_primario, awg_cabo_secundario, secao_condutor_primario, secao_condutor_secundario


def calcular_correntes(potencia_secundaria, tensao_primaria, tensao_secundaria):
    potencia_primaria = potencia_secundaria * 1.1

    corrente_primaria = potencia_primaria / tensao_primaria
    corrente_secundaria = potencia_secundaria / tensao_secundaria

    return corrente_primaria, corrente_secundaria


def calcular_secao_condutor(corrente_primaria, corrente_secundaria, densidade_corrente):
    secao_condutor_primario = corrente_primaria / densidade_corrente
    secao_condutor_secundario = corrente_secundaria / densidade_corrente

    return secao_condutor_primario, secao_condutor_secundario


def calcular_peso_ferro(tipo_lamina, dimensao_a, dimensao_b):
    lamina_info = LaminaInformacoes[tipo_lamina]

    if lamina_info == "Inválido":
        return "Inválido"

    peso_kg_cm = lamina_info[dimensao_a][2]

    return dimensao_b * peso_kg_cm


def calcular_peso_cobre(dimensao_a, dimensao_b, secao_cobre):
    comprimento_espira = 2 * dimensao_b + dimensao_a * (2 + 0.5 * math.pi)

    return ((secao_cobre / 100) * comprimento_espira * 9) / 1000


def calcular_perdas_ferro(frequencia, peso_ferro, inducao_magnetica_max=11300):
    perda_lamina = 1.35 * frequencia * math.pow(inducao_magnetica_max / 10000, 2) / 50

    return 1.15 * perda_lamina * peso_ferro


def calcular_perdas_cobre(densidade_corrente_media, peso_cobre):
    perda_especifica = 2.43 * math.pow(densidade_corrente_media, 2)

    return perda_especifica * peso_cobre


def calcular_rendimento(potencia_secundaria, perdas_nucleo, perdas_cobre):
    return (100 * potencia_secundaria) / (potencia_secundaria + perdas_cobre + perdas_nucleo)


def possibilidade_execucao(tipo_lamina, dimensao_a, espiras_primario, espiras_secundario, secao_condutor_primario,
                           secao_condutor_secundario):
    lamina_info = LaminaInformacoes[tipo_lamina]
    secao_janela = lamina_info[dimensao_a][1]
    sesao_cobre = calcular_secao_cobre(espiras_primario, espiras_secundario, secao_condutor_primario,
                                       secao_condutor_secundario)

    return secao_janela / sesao_cobre >= 3


def plot_corrente_mag(frequencia, espiras_primaria, tensao_primaria, arquivo, plataforma):
    try:
        # Tensão máxima [V]:
        V = tensao_primaria * (2 ** 0.5)
        # Velocidade angular [rad/s]:
        w = 2 * np.pi * frequencia

        # Listas para armazenar os dados
        mmf = dadosMagnetizacao["MMF"]
        fluxo = dadosMagnetizacao["Fluxo"]

        if not plataforma == "android":
            # Carregar o arquivo Excel
            workbook = load_workbook(arquivo)

            # Selecionar a planilha desejada
            sheet = workbook.active  # Ou use workbook['nomedaplanilha'] se souber o nome da planilha

            # Iterar sobre as linhas do arquivo
            for linha in sheet.iter_rows(min_row=2, values_only=True):
                # Adicionar valores às listas, convertendo para int
                mmf.append(linha[0])  # Assumindo que 'MMF' está na primeira coluna
                fluxo.append(linha[1])  # Assumindo que 'Fluxo' está na segunda coluna

            # Fechar o arquivo Excel
            workbook.close()

        t = np.arange(0, 0.034, 1 / 3000)
        fluxo_por_tempo = -V * np.cos(w * t) / (w * espiras_primaria)

        mmf_interpolacao = custom_interpolation(fluxo, mmf, fluxo_por_tempo)

        corrente = mmf_interpolacao / espiras_primaria

        return t, corrente

    except Exception as e:
        print(f"Error: {e}")
